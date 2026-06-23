import json
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESULTS_FILE = os.path.join(os.path.dirname(__file__), 'test-results', 'results.json')
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), 'test-results')

STATUS_COLORS = {
    'passed':   '0070C0',
    'failed':   'C00000',
    'skipped':  'FFC000',
    'timedOut': 'C00000',
}

STATUS_LABELS = {
    'passed':   'PASSED',
    'failed':   'FAILED',
    'skipped':  'SKIPPED',
    'timedOut': 'TIMED OUT',
}

def thin_border():
    s = Side(style='thin', color='D0D0D0')
    return Border(left=s, right=s, top=s, bottom=s)

def load_results(path):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)

def collect_specs(suite, file_title=None):
    """Recursively collect all specs from nested suites."""
    specs = []
    title = file_title or suite.get('title', '')
    for spec in suite.get('specs', []):
        specs.append((title, spec))
    for sub in suite.get('suites', []):
        specs.extend(collect_specs(sub, title))
    return specs

def flatten_tests(data):
    rows = []
    for suite in data.get('suites', []):
        file_title = suite.get('title', '')
        all_specs = collect_specs(suite, file_title)
        for file_t, spec in all_specs:
            spec_title = spec.get('title', '')
            for test in spec.get('tests', []):
                for result in test.get('results', []):
                    status = result.get('status', 'unknown')
                    duration_ms = result.get('duration', 0)
                    error_msg = ''
                    errors = result.get('errors', [])
                    if errors:
                        msg = errors[0].get('message', '')
                        error_msg = msg[:300].replace('\n', ' ') if msg else ''
                    rows.append({
                        'file': file_t,
                        'test': spec_title,
                        'status': status,
                        'duration_s': round(duration_ms / 1000, 2),
                        'error': error_msg,
                    })
    return rows

def write_report(rows, data):
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_path = os.path.join(OUTPUT_DIR, f'test-report_{timestamp}.xlsx')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Test Results'

    total = len(rows)
    passed = sum(1 for r in rows if r['status'] == 'passed')
    failed = sum(1 for r in rows if r['status'] in ('failed', 'timedOut'))
    skipped = sum(1 for r in rows if r['status'] == 'skipped')

    run_at = data.get('stats', {}).get('startTime', datetime.now().isoformat())
    try:
        run_dt = datetime.fromisoformat(run_at.replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M:%S UTC')
    except Exception:
        run_dt = run_at

    summary = [
        ('Run Date', run_dt),
        ('Total Tests', total),
        ('Passed', passed),
        ('Failed', failed),
        ('Skipped', skipped),
        ('Pass Rate', f'{round(passed/total*100, 1)}%' if total else '0%'),
    ]

    label_font = Font(name='Arial', bold=True, size=11)
    value_font = Font(name='Arial', size=11)

    for i, (label, value) in enumerate(summary, start=1):
        lc = ws.cell(row=i, column=1, value=label)
        vc = ws.cell(row=i, column=2, value=value)
        lc.font = label_font
        if label == 'Passed':
            vc.font = Font(name='Arial', bold=True, size=11, color='00B050')
        elif label == 'Failed':
            vc.font = Font(name='Arial', bold=True, size=11, color='C00000')
        else:
            vc.font = value_font

    header_row = len(summary) + 2
    headers = ['Test File', 'Test Name', 'Status', 'Duration (s)', 'Error Message']

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill('solid', start_color='1F3864', end_color='1F3864')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border()

    for row_i, r in enumerate(rows, start=header_row + 1):
        status_label = STATUS_LABELS.get(r['status'], r['status'].upper())
        status_color = STATUS_COLORS.get(r['status'], '808080')

        vals = [r['file'], r['test'], status_label, r['duration_s'], r['error']]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.font = Font(name='Arial', size=10)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical='center', wrap_text=(col == 5))

            if col == 3:
                cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
                cell.fill = PatternFill('solid', start_color=status_color, end_color=status_color)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif row_i % 2 == 0:
                cell.fill = PatternFill('solid', start_color='F2F2F2', end_color='F2F2F2')

    col_widths = [35, 45, 14, 14, 60]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[header_row].height = 22
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    wb.save(out_path)
    return out_path

def main():
    results_path = sys.argv[1] if len(sys.argv) > 1 else RESULTS_FILE
    if not os.path.exists(results_path):
        print(f'ERROR: Results file not found: {results_path}')
        sys.exit(1)

    data = load_results(results_path)
    rows = flatten_tests(data)

    if not rows:
        print('No test results found in JSON.')
        sys.exit(1)

    out = write_report(rows, data)
    print(f'Report generated: {out}')

if __name__ == '__main__':
    main()
